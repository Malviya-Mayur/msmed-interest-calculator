# msmed_calculator/output/exporter.py
"""
Writes detailed and summary DataFrames to a formatted Excel file using openpyxl.
- Two sheets: "Detailed Ledger" and "Summary"
- Bold headers, auto-column widths
- Light red fill (FFCCCC) on rows where Interest Amount > 0
- Indian number formatting for amount columns
- Returns BytesIO for direct FastAPI download
"""

import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Indian number formatting pattern (e.g. 1,00,000.00)
# openpyxl doesn't support Indian locale natively, so we use a custom format string
INDIAN_FORMAT = '#,##,##0.00'
HIGHLIGHT_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

AMOUNT_COLUMNS = {
    "Purchase Amount (₹)", "Amount Settled (₹)", "Interest Amount (₹)",
    "Total Purchases (₹)", "Total Payments (₹)", "Interest-Bearing Amount (₹)",
    "Total Interest Due (₹)"
}


def _write_sheet(ws, df: pd.DataFrame, interest_col: str = None):
    """Write a DataFrame to a worksheet with formatting."""
    headers = list(df.columns)

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 30

    # Write data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        should_highlight = False
        if interest_col and interest_col in df.columns:
            val = row.get(interest_col, 0)
            try:
                should_highlight = float(val) > 0
            except (TypeError, ValueError):
                pass

        for col_idx, header in enumerate(headers, start=1):
            value = row[header]

            # Convert dates/timestamps for Excel compatibility
            if hasattr(value, 'date') and callable(value.date):
                value = value.date()
            elif pd.isna(value) if not isinstance(value, str) else False:
                value = ""

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Apply Indian number format to amount columns
            if header in AMOUNT_COLUMNS:
                cell.number_format = INDIAN_FORMAT

            if should_highlight:
                cell.fill = HIGHLIGHT_FILL

    # Auto-fit column widths
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(header)),
            *[len(str(df.iloc[r, col_idx - 1])) for r in range(min(len(df), 50))]
        ) if len(df) > 0 else len(str(header))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 35)

    # Freeze the header row
    ws.freeze_panes = "A2"


def export_to_excel(detailed_df: pd.DataFrame, summary_df: pd.DataFrame) -> BytesIO:
    """
    Write detailed and summary DataFrames to a BytesIO Excel file.
    Returns the BytesIO object ready to be served as a download.
    """
    wb = Workbook()

    # ── Sheet 1: Detailed Ledger ──────────────────────────────────────────
    ws_detail = wb.active
    ws_detail.title = "Detailed Ledger"
    _write_sheet(ws_detail, detailed_df, interest_col="Interest Amount (₹)")

    # ── Sheet 2: Summary ──────────────────────────────────────────────────
    ws_summary = wb.create_sheet(title="Summary")
    _write_sheet(ws_summary, summary_df, interest_col=None)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
